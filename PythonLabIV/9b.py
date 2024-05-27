import openpyxl

def read(fn , sn):
    wb = openpyxl.load_workbook(fn)
    sheet = wb[sn]

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def write(fn , sn , nd):
    wb = openpyxl.load_workbook(fn)
    sheet = wb[sn]

    for row in nd:
        sheet.append(row)
    wb.save(fn)
    print("Data written to workbook")

inp = 'input.xlsx'
output = 'output.xlsx'

read_data = read(inp , "Sheet1")
for row in read_data:
    print(row)

nd = [
    ("John" , 25),
    ("Harry", 14)
]
write(output , "Sheet1" , nd)